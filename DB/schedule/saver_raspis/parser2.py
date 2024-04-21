import sqlite3
from openpyxl import load_workbook
import re

conn = sqlite3.connect('raspis.db')
cursor = conn.cursor()


#честно не ебу что это тут гпт помог решить трабл с тем что sql выебывается если таблица нчинается с цифр или еще чего то
# функция для создания корректного имени таблицы
def clean_table_name(name):
    # используем регулярное выражение для удаления всех символов кроме букв цифр и подчеркивания
    clean_name = re.sub(r'\W+', '_', name)
    # усли имя начинается с цифры, добавляем "_group_" в начало
    if clean_name[0].isdigit():
        clean_name = f"group_{clean_name}"
    return clean_name


# Функция для создания таблицы с именем
def create_table(group_name):
    group_table_name = clean_table_name(group_name)
    print(f"Creating table: {group_table_name}")
    create_table_query = f"""
    CREATE TABLE IF NOT EXISTS {group_table_name} (
        "Закреплённая кафедра" TEXT,
        "Группа" TEXT,
        "Дисциплина, вид учебной работы" TEXT,
        "Вид занятий" TEXT,
        "Часы" INTEGER,
        "Преподаватель" TEXT,
        "День недели" TEXT,
        "Счет недель" INTEGER,
        "Время" TEXT,
        "Корпус_Аудитория" TEXT
    );
    """
    cursor.execute(create_table_query)
    conn.commit()

# Функция для чтения данных из иксельки и сохранения в бд
def save_to_database(file_path):
    wb = load_workbook(file_path)
    sheet = wb['ОО 2 пг 23-24']
    for row in sheet.iter_rows(min_row=3, values_only=True):  # Начинаем чтение с 3 строки
        group_name = row[1]  # Получаем название группы
        create_table(group_name)  # Создаем таблицу для этой группы если её нет
        new_row = list(row[:9]) + [str(row[9]) + ' ' + str(row[10])]  # Преваращаем в стр перед объединением
        insert_query = f"""
        INSERT INTO {clean_table_name(group_name)} ("Закреплённая кафедра", "Группа", "Дисциплина, вид учебной работы", "Вид занятий", "Часы", "Преподаватель", "День недели", "Счет недель", "Время", "Корпус_Аудитория")
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(insert_query, new_row)
        conn.commit()

save_to_database(r'DB\schedule\saver_raspis\raspisanie_o_itm_22.xlsx')

# Закрываем соединение с бд
conn.close()
