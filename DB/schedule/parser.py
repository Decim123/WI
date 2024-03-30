import sqlite3
import openpyxl
import logging

logging.basicConfig(level=logging.DEBUG)

# Имя файлаи листа с которого нужно извлечь, номер группы и тд все это временно и должен мне это дать Саня
excel_file = 'Расписание весна 2024 г. (2 семестр).xlsx' # название файла(я бы унифицировал как то хз)
sheet_name = '1 КУРС_Журналист_ИТ_Менеджеры' # на какой странице расписание
group_numb = 133 #номер группы(желательно инт)

# нужные переменные
firs_time = 1
header_row = 0
time_column = 0
pn_row= 0
new_header_row = 0
group_row = 0
current_day_of_week = None
days_of_week = ["ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА", "Воскресенье"]
nad_pod = None
group_coloumn = 0
np_column = 0

# Подключение к базе данных
conn = sqlite3.connect('raspis.db')
cursor = conn.cursor()

# Открытие файла
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]

# Создание таблиц в базе данных
cursor.execute('''
    CREATE TABLE IF NOT EXISTS NAD (
        Time TEXT PRIMARY KEY,
        ПОНЕДЕЛЬНИК TEXT,
        ВТОРНИК TEXT,
        СРЕДА TEXT,
        ЧЕТВЕРГ TEXT,
        ПЯТНИЦА TEXT,
        СУББОТА TEXT,
        Воскресен TEXT
    )
''')

cursor.execute('''
    CREATE TABLE IF NOT EXISTS POD (
        Time TEXT PRIMARY KEY,
        ПОНЕДЕЛЬНИК TEXT,
        ВТОРНИК TEXT,
        СРЕДА TEXT,
        ЧЕТВЕРГ TEXT,
        ПЯТНИЦА TEXT,
        СУББОТА TEXT,
        Воскресен TEXT
    )
''')

def saver():
    if current_day_of_week:
            para_time = str(sheet.cell(row=row, column=time_column).value).strip()
            para = str(sheet.cell(row=row, column=group_coloumn).value)
            nad_pod = str(sheet.cell(row=row, column=np_column).value)
            if para_time and para and nad_pod:  # Проверяем что все ключевые значения существуют
                table_name = 'NAD' if nad_pod == 'НАД' else 'POD'
                # Обновляем запись для текущего дня недели на основе времени
                cursor.execute(f"UPDATE {table_name} SET \"{current_day_of_week}\" = ? WHERE Time = ?", (para, para_time))
                logging.debug(f"Запись в таблицу {table_name}: {para} на время {para_time} в {current_day_of_week}")

# Поиск строки с заголовками
for row in range(1, 11):
    if sheet.cell(row=row, column=1).value == 'День недели':
        header_row = row
        break
print(f"Строка с заголовками: {header_row}")

# Проверка на пустую строку под днем недели и поиск номера группы
new_header_row = header_row + 1
if sheet.cell(row=new_header_row, column=1).value == None:
    pn_row = new_header_row + 1
    group_row = header_row
elif sheet.cell(row=new_header_row, column=1).value in days_of_week:
    group_row = new_header_row - 1 
    pn_row = new_header_row
for column in range(1, 20):
    cell_value = sheet.cell(row=group_row, column=column).value
    if cell_value == group_numb:
        group_coloumn = column
for column in range(1, 8):
    if sheet.cell(row=header_row, column=column).value == 'Время':
        time_column = column
for column in range(1, 8):
    if sheet.cell(row=header_row, column=column).value == 'Над/Под':
        np_column = column
print(pn_row, group_coloumn, time_column, np_column)


unique_times = set()  # Множество для хранения уникальных времен

for row in range(pn_row, sheet.max_row + 1):
    para_time = str(sheet.cell(row=row, column=time_column).value).strip()
    # Проверяем что время не пустое и строка содержит данные расписания
    if para_time and any(sheet.cell(row=row, column=col).value for col in range(2, sheet.max_column + 1)):
        if para_time not in unique_times:
            unique_times.add(para_time)
            # Для каждого уникального времени проверяем существует ли запись и если нет то создаем
            cursor.execute(f"INSERT INTO NAD (Time) VALUES (?) ON CONFLICT(Time) DO NOTHING", (para_time,))
            cursor.execute(f"INSERT INTO POD (Time) VALUES (?) ON CONFLICT(Time) DO NOTHING", (para_time,))

# когда у нас есть записи для каждого времени обновляем данные для каждого дня недели
for row in range(pn_row, sheet.max_row + 1):
    day_of_week_cell_value = sheet.cell(row=row, column=1).value
    if day_of_week_cell_value in days_of_week:
        current_day_of_week = day_of_week_cell_value
        saver()
    else:
        saver()

# Переносим подтверждение изменений и закрытие соединения
conn.commit()
conn.close()
